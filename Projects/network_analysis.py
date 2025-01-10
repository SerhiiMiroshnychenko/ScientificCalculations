import networkx as nx
import logging
from typing import Dict, List, Tuple, Set
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import csv

# Налаштування логування
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('network_analysis.log', mode='w', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class NetworkAnalyzer:
    def __init__(self):
        """Ініціалізація графа та базових даних"""
        self.graph = nx.DiGraph()
        # Оновлюємо тривалості робіт відповідно до графіка
        durations = {
            (0, 1): 31.98,  # Розробка технічних умов
            (1, 2): 34.12,  # Загальне компонування виробу
            (1, 7): 38.38,  # Видача завдання на документацію
            (2, 3): 38.38,  # Проєктування електричного блоку
            (2, 4): 38.38,  # Проєктування механічного блоку
            (2, 5): 51.17,  # Оформлення замовлень на компоненти
            (3, 6): 42.65,  # Виготовлення електричної частини
            (4, 6): 42.65,  # Виготовлення механічної частини
            (5, 6): 44.78,  # Отримання компонентів
            (6, 7): 38.38,  # Підготовка технічної інформації
            (7, 8): 34.12,  # Складання виробу
            (7, 9): 25.59,  # Розробка експлуатаційної документації
            (8, 9): 25.59,  # Контрольні випробування
        }
        for edge, duration in durations.items():
            self.graph.add_edge(edge[0], edge[1], weight=duration)

        logging.info("Граф ініціалізовано з %d вершинами та %d ребрами",
                     self.graph.number_of_nodes(),
                     self.graph.number_of_edges())

    def find_all_paths(self) -> Dict[str, Tuple[List[int], int]]:
        """Знаходження всіх можливих шляхів від початкової до кінцевої вершини"""
        logging.info("Пошук всіх можливих шляхів від вершини 0 до 9")
        paths = {}
        path_number = 1

        def calculate_path_length(path: List[int]) -> int:
            return sum(self.graph[path[i]][path[i + 1]]['weight'] for i in range(len(path) - 1))

        for path in nx.all_simple_paths(self.graph, 0, 9):
            length = calculate_path_length(path)
            path_name = f"L{path_number}"
            paths[path_name] = (path, length)

            logging.info(f"Знайдено шлях {path_name}: {' -> '.join(map(str, path))}")
            logging.info(f"Тривалість шляху {path_name}: {length} тижнів")

            path_number += 1

        return paths

    def find_critical_path(self, paths: Dict[str, Tuple[List[int], int]]) -> Tuple[List[int], int]:
        """Визначення критичного шляху"""
        critical_path = max(paths.items(), key=lambda x: x[1][1])
        logging.info("Критичний шлях %s: %s",
                     critical_path[0],
                     ' -> '.join(map(str, critical_path[1][0])))
        logging.info("Тривалість критичного шляху: %d тижнів", critical_path[1][1])
        return critical_path[1]

    def calculate_earliest_times(self) -> Dict[int, int]:
        """Розрахунок найраніших термінів завершення подій"""
        logging.info("Розрахунок найраніших термінів завершення подій")
        earliest_times = {node: 0 for node in self.graph.nodes()}

        for node in nx.topological_sort(self.graph):
            for pred in self.graph.predecessors(node):
                earliest_times[node] = max(
                    earliest_times[node],
                    earliest_times[pred] + self.graph[pred][node]['weight']
                )
            logging.info(f"Тр{node} = {earliest_times[node]}")

        return earliest_times

    def calculate_latest_times(self, critical_length: int) -> Dict[int, int]:
        """Розрахунок найпізніших термінів завершення подій"""
        logging.info("Розрахунок найпізніших термінів завершення подій")
        latest_times = {node: critical_length for node in self.graph.nodes()}

        for node in reversed(list(nx.topological_sort(self.graph))):
            if list(self.graph.successors(node)):
                latest_times[node] = min(
                    latest_times[succ] - self.graph[node][succ]['weight']
                    for succ in self.graph.successors(node)
                )
            logging.info(f"Тп{node} = {latest_times[node]}")

        return latest_times

    def calculate_time_reserves(self, earliest_times: Dict[int, int],
                                latest_times: Dict[int, int]) -> Dict[int, int]:
        """Розрахунок часових резервів подій"""
        logging.info("Розрахунок часових резервів подій")
        reserves = {}

        for node in self.graph.nodes():
            reserves[node] = latest_times[node] - earliest_times[node]
            logging.info(f"R{node} = Тп{node} - Тр{node} = {latest_times[node]} - {earliest_times[node]} = {reserves[node]}")

        return reserves

    def calculate_work_parameters(self, earliest_times: Dict[int, int],
                                  latest_times: Dict[int, int]) -> Dict[Tuple[int, int], Dict]:
        """Розрахунок параметрів робіт"""
        logging.info("Розрахунок параметрів робіт")
        work_params = {}

        for start, end in self.graph.edges():
            duration = self.graph[start][end]['weight']

            # Найраніший час початку
            earliest_start = earliest_times[start]
            # Найраніший час завершення
            earliest_finish = earliest_start + duration
            # Найпізніший час завершення
            latest_finish = latest_times[end]
            # Найпізніший час початку
            latest_start = latest_finish - duration

            # Повний резерв часу
            total_reserve = latest_finish - earliest_times[start] - duration

            # Частковий резерв першого типу
            first_type_reserve = latest_times[end] - earliest_times[start] - duration

            # Частковий резерв другого типу
            second_type_reserve = earliest_times[end] - earliest_times[start] - duration

            work_params[(start, end)] = {
                'duration': duration,
                'earliest_start': earliest_start,
                'earliest_finish': earliest_finish,
                'latest_start': latest_start,
                'latest_finish': latest_finish,
                'total_reserve': total_reserve,
                'first_type_reserve': first_type_reserve,
                'second_type_reserve': second_type_reserve
            }

            logging.info(f"\nРобота ({start},{end}):")
            logging.info(f"Тривалість: {duration}")
            logging.info(f"Найраніший початок: {earliest_start}")
            logging.info(f"Найраніше завершення: {earliest_finish}")
            logging.info(f"Найпізніший початок: {latest_start}")
            logging.info(f"Найпізніше завершення: {latest_finish}")
            logging.info(f"Повний резерв: {total_reserve}")
            logging.info(f"Частковий резерв 1-го типу: {first_type_reserve}")
            logging.info(f"Частковий резерв 2-го типу: {second_type_reserve}")

        return work_params

    def calculate_tension_coefficients(self, critical_path: List[int],
                                       critical_length: int) -> Dict[Tuple[int, int], float]:
        """
        Розрахунок коефіцієнтів напруженості робіт за формулою:
        Kij = (t[Lmax] - t1[Lкр]) / (t[Lкр] - t1[Lкр])
        де:
        t[Lmax] – тривалість максимального шляху через роботу
        t[Lкр] – тривалість критичного шляху
        t1[Lкр] – тривалість спільного відрізку максимального шляху з критичним
        """
        logging.info("Розрахунок коефіцієнтів напруженості робіт")
        tension_coeffs = {}
        critical_edges = set(zip(critical_path[:-1], critical_path[1:]))

        for start, end in self.graph.edges():
            if (start, end) in critical_edges:
                tension_coeffs[(start, end)] = 1.0
                logging.info(f"Кн({start},{end}) = 1.0 (критичний шлях)")
                continue

            # Знаходження максимального шляху через роботу
            max_path_length = 0
            max_path = None
            for path in nx.all_simple_paths(self.graph, 0, 9):
                if start in path and end in path and path.index(end) == path.index(start) + 1:
                    length = sum(self.graph[path[i]][path[i + 1]]['weight']
                                 for i in range(len(path) - 1))
                    if length > max_path_length:
                        max_path_length = length
                        max_path = path

            if max_path is None:
                tension_coeffs[(start, end)] = 0.0
                logging.info(f"Кн({start},{end}) = 0.0 (немає шляху)")
                continue

            # Знаходження спільного відрізку з критичним шляхом
            common_length = 0
            current_common = 0
            i = 0
            while i < len(max_path) - 1:
                if i < len(max_path) - 1 and (max_path[i], max_path[i + 1]) in critical_edges:
                    current_common += self.graph[max_path[i]][max_path[i + 1]]['weight']
                else:
                    common_length = max(common_length, current_common)
                    current_common = 0
                i += 1
            common_length = max(common_length, current_common)

            # Розрахунок коефіцієнта напруженості
            if critical_length != common_length:
                tension = (max_path_length - common_length) / (critical_length - common_length)
                tension_coeffs[(start, end)] = round(tension, 4)  # Збільшуємо точність до 4 знаків
                logging.info(f"Кн({start},{end}) = {tension_coeffs[(start, end)]} "
                             f"(t[Lmax]={max_path_length}, t1[Lкр]={common_length}, t[Lкр]={critical_length})")
            else:
                tension_coeffs[(start, end)] = 1.0
                logging.info(f"Кн({start},{end}) = 1.0")

        return tension_coeffs

    def export_to_excel(self, filename: str, work_params: Dict[Tuple[int, int], Dict],
                        earliest_times: Dict[int, int], latest_times: Dict[int, int],
                        reserves: Dict[int, int], tension_coeffs: Dict[Tuple[int, int], float]):
        """Експорт результатів обчислень у Excel файл"""
        logging.info(f"Експорт результатів у файл {filename}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Результати обчислень"

        # Налаштування стилів
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center')

        # Заголовки колонок
        headers = ['i', 'Tpi', 'Tпi', 'Ri', 'i', 'j', 'tij', 'Tрпij', 'Tрзij',
                   'Tппij', 'Tпзij', 'Rпij', 'Rч1ij', 'Rч2ij', 'Kнij']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        # Заповнення даних
        row = 2
        for node in sorted(self.graph.nodes()):
            # Якщо є вихідні ребра
            if list(self.graph.successors(node)):
                for succ in self.graph.successors(node):
                    ws.cell(row=row, column=1, value=node)
                    ws.cell(row=row, column=2, value=earliest_times[node])
                    ws.cell(row=row, column=3, value=latest_times[node])
                    ws.cell(row=row, column=4, value=reserves[node])
                    ws.cell(row=row, column=5, value=node)
                    ws.cell(row=row, column=6, value=succ)
                    ws.cell(row=row, column=7, value=self.graph[node][succ]['weight'])

                    params = work_params[(node, succ)]
                    ws.cell(row=row, column=8, value=params['earliest_start'])
                    ws.cell(row=row, column=9, value=params['earliest_finish'])
                    ws.cell(row=row, column=10, value=params['latest_start'])
                    ws.cell(row=row, column=11, value=params['latest_finish'])
                    ws.cell(row=row, column=12, value=params['total_reserve'])
                    ws.cell(row=row, column=13, value=params['first_type_reserve'])
                    ws.cell(row=row, column=14, value=params['second_type_reserve'])
                    ws.cell(row=row, column=15, value=tension_coeffs[(node, succ)])

                    # Застосування стилів
                    for col in range(1, 16):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = center_alignment

                    row += 1
            else:
                # Для кінцевої вершини
                ws.cell(row=row, column=1, value=node)
                ws.cell(row=row, column=2, value=earliest_times[node])
                ws.cell(row=row, column=3, value=latest_times[node])
                ws.cell(row=row, column=4, value=reserves[node])
                ws.cell(row=row, column=5, value='-')
                ws.cell(row=row, column=6, value='-')
                for col in range(7, 16):
                    ws.cell(row=row, column=col, value='-')

                # Застосування стилів
                for col in range(1, 16):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = center_alignment

                row += 1

        # Автоматичне налаштування ширини колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Збереження файлу
        wb.save(filename)
        logging.info(f"Результати успішно експортовано у файл {filename}")

    def export_to_csv(self, filename: str, work_params: Dict[Tuple[int, int], Dict],
                      earliest_times: Dict[int, int], latest_times: Dict[int, int],
                      reserves: Dict[int, int], tension_coeffs: Dict[Tuple[int, int], float]):
        """Експорт результатів обчислень у CSV файл"""
        logging.info(f"Експорт результатів у файл {filename}")

        # Заголовки колонок
        headers = ['i', 'Tpi', 'Tпi', 'Ri', 'i', 'j', 'tij', 'Tрпij', 'Tрзij',
                   'Tппij', 'Tпзij', 'Rпij', 'Rч1ij', 'Rч2ij', 'Kнij']

        def format_value(value):
            """Форматування значень для CSV"""
            if isinstance(value, (int, float)):
                return f"{value:.2f}"
            return value

        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(headers)

            # Заповнення даних
            for node in sorted(self.graph.nodes()):
                # Якщо є вихідні ребра
                if list(self.graph.successors(node)):
                    for succ in self.graph.successors(node):
                        row_data = [
                            format_value(node),
                            format_value(earliest_times[node]),
                            format_value(latest_times[node]),
                            format_value(reserves[node]),
                            format_value(node),
                            format_value(succ),
                            format_value(self.graph[node][succ]['weight']),
                            format_value(work_params[(node, succ)]['earliest_start']),
                            format_value(work_params[(node, succ)]['earliest_finish']),
                            format_value(work_params[(node, succ)]['latest_start']),
                            format_value(work_params[(node, succ)]['latest_finish']),
                            format_value(work_params[(node, succ)]['total_reserve']),
                            format_value(work_params[(node, succ)]['first_type_reserve']),
                            format_value(work_params[(node, succ)]['second_type_reserve']),
                            format_value(tension_coeffs[(node, succ)])
                        ]
                        writer.writerow(row_data)
                else:
                    # Для кінцевої вершини
                    row_data = [
                        format_value(node),
                        format_value(earliest_times[node]),
                        format_value(latest_times[node]),
                        format_value(reserves[node]),
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-',
                        '-'
                    ]
                    writer.writerow(row_data)

        logging.info(f"Результати успішно експортовано у файл {filename}")

    def analyze_network(self):
        """Повний аналіз сіткового графіка"""
        logging.info("Початок аналізу сіткового графіка")

        # Знаходження всіх шляхів
        paths = self.find_all_paths()

        # Визначення критичного шляху
        critical_path, critical_length = self.find_critical_path(paths)

        # Розрахунок найраніших термінів
        earliest_times = self.calculate_earliest_times()

        # Розрахунок найпізніших термінів
        latest_times = self.calculate_latest_times(critical_length)

        # Розрахунок резервів часу
        reserves = self.calculate_time_reserves(earliest_times, latest_times)

        # Розрахунок параметрів робіт
        work_params = self.calculate_work_parameters(earliest_times, latest_times)

        # Розрахунок коефіцієнтів напруженості
        tension_coeffs = self.calculate_tension_coefficients(critical_path, critical_length)

        # Експорт результатів у Excel та CSV
        self.export_to_excel('network_analysis_results.xlsx', work_params,
                             earliest_times, latest_times, reserves, tension_coeffs)
        self.export_to_csv('network_analysis_results.csv', work_params,
                           earliest_times, latest_times, reserves, tension_coeffs)

        return {
            'paths': paths,
            'critical_path': critical_path,
            'critical_length': critical_length,
            'earliest_times': earliest_times,
            'latest_times': latest_times,
            'reserves': reserves,
            'work_params': work_params,
            'tension_coeffs': tension_coeffs
        }

if __name__ == "__main__":
    analyzer = NetworkAnalyzer()
    results = analyzer.analyze_network()
